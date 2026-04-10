"""
Services para o app reports — exportação de dados e agendamento de relatórios.
"""
import csv
from datetime import timedelta
from io import BytesIO

from dateutil.relativedelta import relativedelta
from django.core.mail import send_mail, EmailMessage
from django.http import HttpResponse
from django.utils import timezone

from apps.reports.models import ScheduledReport
from apps.reports.selectors import get_filtered_transactions


# ---------------------------------------------------------------------------
# Exportação
# ---------------------------------------------------------------------------

CABECALHOS = ["Nome", "Valor", "Tipo", "Categoria", "Cartão", "Data", "Status"]


def _transaction_row(tx):
    """Retorna uma lista com os campos de uma transação para exportação."""
    return [
        tx.name,
        str(tx.amount),
        tx.get_type_display(),
        tx.category.name if tx.category else "",
        tx.card.name if tx.card else "",
        tx.date.strftime("%d/%m/%Y") if tx.date else "",
        tx.get_status_display(),
    ]


def export_csv(transactions_qs):
    """Gera arquivo CSV e retorna como HttpResponse para download."""
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="relatorio.csv"'

    writer = csv.writer(response)
    writer.writerow(CABECALHOS)

    for tx in transactions_qs.select_related("category", "card"):
        writer.writerow(_transaction_row(tx))

    return response


def export_xlsx(transactions_qs):
    """Gera arquivo Excel (openpyxl) e retorna como HttpResponse para download."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"

    # Cabeçalhos em negrito
    for col, header in enumerate(CABECALHOS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Dados
    for row_idx, tx in enumerate(transactions_qs.select_related("category", "card"), 2):
        for col_idx, value in enumerate(_transaction_row(tx), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="relatorio.xlsx"'
    return response


def export_pdf(transactions_qs, user, filters):
    """Gera relatório PDF formatado usando weasyprint e retorna como HttpResponse."""
    import weasyprint
    from django.db.models import Sum
    from decimal import Decimal

    total_entradas = transactions_qs.filter(type="entrada").aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0")
    total_saidas = transactions_qs.filter(type="saida").aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0")
    saldo = total_entradas - total_saidas

    # Monta HTML do relatório
    rows_html = ""
    for tx in transactions_qs.select_related("category", "card"):
        row = _transaction_row(tx)
        rows_html += "<tr>" + "".join(f"<td>{v}</td>" for v in row) + "</tr>"

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{ font-family: Arial, sans-serif; font-size: 12px; margin: 20px; }}
            h1 {{ color: #333; font-size: 18px; }}
            .resumo {{ margin-bottom: 20px; }}
            .resumo span {{ margin-right: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            th, td {{ border: 1px solid #ddd; padding: 6px 8px; text-align: left; }}
            th {{ background-color: #f5f5f5; font-weight: bold; }}
        </style>
    </head>
    <body>
        <h1>Relatório Financeiro — {user.email}</h1>
        <div class="resumo">
            <span><b>Entradas:</b> R$ {total_entradas}</span>
            <span><b>Saídas:</b> R$ {total_saidas}</span>
            <span><b>Saldo:</b> R$ {saldo}</span>
        </div>
        <table>
            <thead>
                <tr>{"".join(f"<th>{h}</th>" for h in CABECALHOS)}</tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>
    </body>
    </html>
    """

    pdf_bytes = weasyprint.HTML(string=html).write_pdf()

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="relatorio.pdf"'
    return response


# ---------------------------------------------------------------------------
# Agendamento
# ---------------------------------------------------------------------------

def _calcular_next_send_at(frequency):
    """Calcula a próxima data de envio com base na frequência."""
    agora = timezone.now()
    if frequency == "semanal":
        return agora + timedelta(days=7)
    elif frequency == "quinzenal":
        return agora + timedelta(days=14)
    elif frequency == "mensal":
        return agora + relativedelta(months=1)
    return agora + timedelta(days=7)


def create_scheduled_report(user, name, frequency, export_format, filters):
    """Cria um relatório agendado com next_send_at calculado automaticamente."""
    return ScheduledReport.objects.create(
        user=user,
        name=name,
        frequency=frequency,
        export_format=export_format,
        filters=filters,
        next_send_at=_calcular_next_send_at(frequency),
    )


def update_scheduled_report(report_id, user, **kwargs):
    """Atualiza configuração de um relatório agendado. Recalcula next_send_at se frequência alterada."""
    try:
        report = ScheduledReport.objects.get(id=report_id, is_active=True)
    except ScheduledReport.DoesNotExist:
        raise PermissionError("Relatório agendado não encontrado.")

    if report.user != user:
        raise PermissionError("Sem permissão para editar este relatório.")

    for campo, valor in kwargs.items():
        setattr(report, campo, valor)

    # Recalcula next_send_at se frequência foi alterada
    if "frequency" in kwargs:
        report.next_send_at = _calcular_next_send_at(kwargs["frequency"])

    report.save()
    return report


def deactivate_scheduled_report(report_id, user):
    """Soft-delete de um relatório agendado."""
    try:
        report = ScheduledReport.objects.get(id=report_id, is_active=True)
    except ScheduledReport.DoesNotExist:
        raise PermissionError("Relatório agendado não encontrado.")

    if report.user != user:
        raise PermissionError("Sem permissão para desativar este relatório.")

    report.is_active = False
    report.save()
    return report


def process_due_reports():
    """
    Processa relatórios agendados com next_send_at <= agora e is_active=True.
    Para cada: gera o arquivo, envia por e-mail, atualiza datas.
    """
    agora = timezone.now()
    due_reports = ScheduledReport.objects.filter(
        next_send_at__lte=agora,
        is_active=True,
    )

    for report in due_reports:
        # Gera o queryset filtrado
        qs = get_filtered_transactions(report.user, report.filters)

        # Gera o arquivo no formato configurado
        if report.export_format == "csv":
            response = export_csv(qs)
            filename = "relatorio.csv"
            content_type = "text/csv"
        elif report.export_format == "xlsx":
            response = export_xlsx(qs)
            filename = "relatorio.xlsx"
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            response = export_pdf(qs, report.user, report.filters)
            filename = "relatorio.pdf"
            content_type = "application/pdf"

        # Envia por e-mail com anexo
        email = EmailMessage(
            subject=f"Relatório Agendado: {report.name}",
            body=f"Segue em anexo o relatório '{report.name}' gerado automaticamente.",
            to=[report.user.email],
        )
        email.attach(filename, response.content, content_type)
        email.send()

        # Atualiza datas
        report.last_sent_at = agora
        report.next_send_at = _calcular_next_send_at(report.frequency)
        report.save()
