"""
Testes unitários para services do app reports — exportação e agendamento.
"""
import pytest
from datetime import date, timedelta
from decimal import Decimal
from unittest.mock import patch
from io import BytesIO

from django.utils import timezone

from apps.accounts.models import CustomUser
from apps.cards.models import Card
from apps.categories.models import Category
from apps.transactions.models import Transaction
from apps.reports.models import ScheduledReport


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def user(db):
    return CustomUser.objects.create_user(
        email="export@test.com",
        nickname="exportuser",
        password="pass12345",
    )


@pytest.fixture
def category(user):
    return Category.objects.create(user=user, name="Alimentação", is_active=True)


@pytest.fixture
def card(user):
    return Card.objects.create(
        user=user, name="Nubank", brand="Mastercard",
        last_four_digits="1234", card_type="credito",
        credit_limit=Decimal("5000.00"),
    )


@pytest.fixture
def transactions(user, category, card):
    return [
        Transaction.objects.create(
            user=user, name="Salário", amount=Decimal("5000.00"),
            type="entrada", status="pago", category=category,
            date=date(2024, 1, 15),
        ),
        Transaction.objects.create(
            user=user, name="Almoço", amount=Decimal("50.00"),
            type="saida", status="pago", category=category, card=card,
            date=date(2024, 1, 20),
        ),
    ]


@pytest.fixture
def transactions_qs(user, transactions):
    return Transaction.objects.filter(user=user, is_active=True)


# ---------------------------------------------------------------------------
# Testes: export_csv
# ---------------------------------------------------------------------------

class TestExportCSV:
    def test_retorna_csv_valido(self, transactions_qs):
        from apps.reports.services import export_csv
        response = export_csv(transactions_qs)

        assert response["Content-Type"] == "text/csv"
        assert "attachment" in response["Content-Disposition"]
        content = response.content.decode("utf-8")
        lines = content.strip().split("\n")
        # Cabeçalho + 2 linhas de dados
        assert len(lines) == 3

    def test_csv_cabecalhos(self, transactions_qs):
        from apps.reports.services import export_csv
        response = export_csv(transactions_qs)
        content = response.content.decode("utf-8")
        header = content.split("\n")[0].strip()
        assert "Nome" in header
        assert "Valor" in header
        assert "Tipo" in header

    def test_csv_sem_dados(self, user):
        from apps.reports.services import export_csv
        qs = Transaction.objects.filter(user=user, is_active=True)
        response = export_csv(qs)
        content = response.content.decode("utf-8")
        lines = content.strip().split("\n")
        # Apenas cabeçalho
        assert len(lines) == 1


# ---------------------------------------------------------------------------
# Testes: export_xlsx
# ---------------------------------------------------------------------------

class TestExportXLSX:
    def test_retorna_xlsx_valido(self, transactions_qs):
        from apps.reports.services import export_xlsx
        response = export_xlsx(transactions_qs)

        assert response["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in response["Content-Disposition"]
        # Verifica que o conteúdo é um arquivo xlsx válido (openpyxl)
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        assert ws.max_row == 3  # header + 2 dados

    def test_xlsx_cabecalhos(self, transactions_qs):
        from apps.reports.services import export_xlsx
        import openpyxl
        response = export_xlsx(transactions_qs)
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "Nome" in headers
        assert "Valor" in headers


# ---------------------------------------------------------------------------
# Testes: export_pdf
# ---------------------------------------------------------------------------

class TestExportPDF:
    def test_retorna_pdf_valido(self, transactions_qs, user):
        from apps.reports.services import export_pdf
        response = export_pdf(transactions_qs, user, {})

        assert response["Content-Type"] == "application/pdf"
        assert "attachment" in response["Content-Disposition"]
        # PDF começa com %PDF
        assert response.content[:5] == b"%PDF-"


# ---------------------------------------------------------------------------
# Testes: create_scheduled_report
# ---------------------------------------------------------------------------

class TestCreateScheduledReport:
    def test_cria_relatorio_semanal(self, user):
        from apps.reports.services import create_scheduled_report
        report = create_scheduled_report(
            user=user, name="Semanal CSV",
            frequency="semanal", export_format="csv",
            filters={"type": "saida"},
        )
        assert report.name == "Semanal CSV"
        assert report.frequency == "semanal"
        assert report.export_format == "csv"
        # next_send_at deve ser ~7 dias no futuro
        diff = report.next_send_at - timezone.now()
        assert 6 <= diff.days <= 7

    def test_cria_relatorio_quinzenal(self, user):
        from apps.reports.services import create_scheduled_report
        report = create_scheduled_report(
            user=user, name="Quinzenal",
            frequency="quinzenal", export_format="xlsx",
            filters={},
        )
        diff = report.next_send_at - timezone.now()
        assert 13 <= diff.days <= 14

    def test_cria_relatorio_mensal(self, user):
        from apps.reports.services import create_scheduled_report
        report = create_scheduled_report(
            user=user, name="Mensal",
            frequency="mensal", export_format="pdf",
            filters={},
        )
        diff = report.next_send_at - timezone.now()
        assert 27 <= diff.days <= 31


# ---------------------------------------------------------------------------
# Testes: update_scheduled_report
# ---------------------------------------------------------------------------

class TestUpdateScheduledReport:
    def test_atualiza_nome(self, user):
        from apps.reports.services import create_scheduled_report, update_scheduled_report
        report = create_scheduled_report(
            user=user, name="Original",
            frequency="semanal", export_format="csv", filters={},
        )
        updated = update_scheduled_report(report.id, user, name="Atualizado")
        assert updated.name == "Atualizado"

    def test_recalcula_next_send_at_ao_mudar_frequencia(self, user):
        from apps.reports.services import create_scheduled_report, update_scheduled_report
        report = create_scheduled_report(
            user=user, name="Teste",
            frequency="semanal", export_format="csv", filters={},
        )
        old_next = report.next_send_at
        updated = update_scheduled_report(report.id, user, frequency="mensal")
        assert updated.next_send_at != old_next

    def test_tenant_isolation(self, user):
        from apps.reports.services import create_scheduled_report, update_scheduled_report
        other = CustomUser.objects.create_user(
            email="intruder@test.com", nickname="intruder", password="pass12345",
        )
        report = create_scheduled_report(
            user=user, name="Meu", frequency="semanal",
            export_format="csv", filters={},
        )
        with pytest.raises(PermissionError):
            update_scheduled_report(report.id, other, name="Hack")


# ---------------------------------------------------------------------------
# Testes: deactivate_scheduled_report
# ---------------------------------------------------------------------------

class TestDeactivateScheduledReport:
    def test_soft_delete(self, user):
        from apps.reports.services import create_scheduled_report, deactivate_scheduled_report
        report = create_scheduled_report(
            user=user, name="Deletar",
            frequency="semanal", export_format="csv", filters={},
        )
        result = deactivate_scheduled_report(report.id, user)
        assert result.is_active is False

    def test_tenant_isolation(self, user):
        from apps.reports.services import create_scheduled_report, deactivate_scheduled_report
        other = CustomUser.objects.create_user(
            email="intruder2@test.com", nickname="intruder2", password="pass12345",
        )
        report = create_scheduled_report(
            user=user, name="Meu", frequency="semanal",
            export_format="csv", filters={},
        )
        with pytest.raises(PermissionError):
            deactivate_scheduled_report(report.id, other)


# ---------------------------------------------------------------------------
# Testes: process_due_reports
# ---------------------------------------------------------------------------

class TestProcessDueReports:
    @patch("apps.reports.services.EmailMessage")
    def test_processa_relatorio_pendente(self, mock_email_cls, user, transactions):
        from apps.reports.services import process_due_reports
        report = ScheduledReport.objects.create(
            user=user, name="Pendente", frequency="semanal",
            export_format="csv", filters={},
            next_send_at=timezone.now() - timedelta(hours=1),
        )
        process_due_reports()

        mock_email_cls.assert_called_once()
        mock_email_cls.return_value.send.assert_called_once()
        report.refresh_from_db()
        assert report.last_sent_at is not None
        # next_send_at deve ser ~7 dias no futuro
        diff = report.next_send_at - timezone.now()
        assert 6 <= diff.days <= 7

    @patch("apps.reports.services.EmailMessage")
    def test_ignora_relatorio_inativo(self, mock_email_cls, user):
        from apps.reports.services import process_due_reports
        ScheduledReport.objects.create(
            user=user, name="Inativo", frequency="semanal",
            export_format="csv", filters={},
            next_send_at=timezone.now() - timedelta(hours=1),
            is_active=False,
        )
        process_due_reports()
        mock_email_cls.assert_not_called()

    @patch("apps.reports.services.EmailMessage")
    def test_ignora_relatorio_futuro(self, mock_email_cls, user):
        from apps.reports.services import process_due_reports
        ScheduledReport.objects.create(
            user=user, name="Futuro", frequency="semanal",
            export_format="csv", filters={},
            next_send_at=timezone.now() + timedelta(days=3),
        )
        process_due_reports()
        mock_email_cls.assert_not_called()
